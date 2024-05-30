from openpyxl import Workbook, load_workbook


def save_excel(df, file_name):
    """
    Recebe arquivo excel e ajusta o tamanho das colunas ao conteúdo

    Args:
        excel_file (str): Caminho e nome do arquivo excel

    Returns:
        None
    """
    excel_file = f"{file_name}.xlsx"
    
    try:
        df_final = df.drop_duplicates(subset=['cnpj'])
        df_final.to_excel(excel_file, index=False, engine='openpyxl')
        format_excel(excel_file)
    except PermissionError as e:
        print("Feche o arquivo excel e tente novamente!")

def format_excel(excel_file):
    """
    Recebe arquivo excel e ajusta o tamanho das colunas ao conteúdo

    Args:
        excel_file (str): Caminho e nome do arquivo excel

    Returns:
        None
    """
    try:
        wb = Workbook()
        wb = load_workbook(excel_file)
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        wb.save(excel_file)
    except PermissionError as e:
        print("Feche o arquivo excel e tente novamente!")