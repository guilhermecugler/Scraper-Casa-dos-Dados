#encoding: utf-8
from urllib.request import Request, urlopen, urlretrieve
from urllib.error import URLError, HTTPError
from bs4 import BeautifulSoup
import pandas as pd
import json
import requests
import io
from sheet2dict import Worksheet
import os, sys
import math
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment
import win32com.client as client
import datetime as dt
import smtplib, ssl
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage

import importlib.util

MODULE_PATH = "mensagem.py"
MODULE_NAME = "mensagem"

def Convert(lst):
    res_dct = {lst[i]: lst[i + 1] for i in range(0, len(lst), 2)}
    return res_dct

def adicionarPlanilha(nomePlanilha, df, enviar_email):

    isFile = os.path.isfile(nomePlanilha)
    if(isFile == False):
        writer = pd.ExcelWriter(nomePlanilha, engine='xlsxwriter')
        writer.close()

    ws = Worksheet()
    
    dictPlanilha = ws.xlsx_to_dict(path=nomePlanilha)
    dfPlanilha = pd.DataFrame(data=dictPlanilha.sheet_items)
    dfFinal1 = pd.concat(df)
    dfFinal = pd.concat([dfPlanilha, dfFinal1])
    
    if enviar_email:
        enviar="Sim"
    else:
        enviar = "Não"

    if 'Enviar e-mail?' in dfPlanilha.columns:
        dfFinal['Enviar e-mail?'] = enviar
    else:
        dfFinal.insert(0, 'Enviar e-mail?', enviar)





    try:
        dfFinal.to_excel(nomePlanilha, index=False)
        cnpj = dfFinal

        writer = pd.ExcelWriter(nomePlanilha, engine='openpyxl', if_sheet_exists='replace', mode='a')

        workbook  = writer.book
        std=workbook.get_sheet_names()
        sheet = workbook.get_sheet_by_name('Sheet1')
        for column_cells in sheet.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = new_column_length*1.23

        for row in range(1,sheet.max_row+1):
            for col in range(1,sheet.max_column+1):
                cell=sheet.cell(row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        
        writer.close()
        return "ok"
    except PermissionError as pe:
        return pe.errno

def buscarEmpresas(json_data):
        dados_json = {}
        quantidadeCNPJ = []

        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36'}
        api_search = "https://api.casadosdados.com.br/v2/public/cnpj/search"

        try:

            response = requests.post(api_search, headers=headers, json=json_data)

        except HTTPError as e:
            print(e.status, e.reason)

        except URLError as e:
            print(e.reason)
        
        dados_json = json.loads(response.content)
        if(dados_json['success'] == False):
            print("Erro ao pegar dados: \n"+dados_json['message'])
            raise Exception("Erro ao pegar dados")

        quantidadeCNPJ = dados_json['data']['count']

        if quantidadeCNPJ == 0:
            raise Exception("Nenhum dado encontrado")

        if quantidadeCNPJ > 1000:
             quantidade_paginas = 50
        else:
            quantidade_paginas = math.ceil(quantidadeCNPJ / 20)

        return json_data, quantidade_paginas, quantidadeCNPJ

def enviarEmail(assinatura):
    
    spec = importlib.util.spec_from_file_location(MODULE_NAME, MODULE_PATH)
    mensagem = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mensagem
    spec.loader.exec_module(mensagem)
    
    tabela = pd.read_excel('Resultados.xlsx')

    tabela_enviar = tabela.loc[tabela['Enviar e-mail?']=='Sim']
    tabela_enviar = tabela_enviar.loc[tabela_enviar['Enviar e-mail?'] == 'Sim']

    dados= tabela_enviar.values.tolist()

    for dado in dados:
        destinatario = dado[18]
        nome=dado[2]
        cnpj=dado[1]
        nome_empresa=dado[3]
        port = 587  # For starttls
        smtp_server = "smtp-mail.outlook.com"
        sender_email = mensagem.email
        password = mensagem.senha

        email_message = MIMEMultipart()
        email_message['From'] = sender_email
        email_message['To'] = destinatario
        email_message['Subject'] = mensagem.assunto

        message = mensagem.mensagem

        assinatura2 = mensagem.assinatura

        email_message.attach(MIMEText(message, "html"))

        if assinatura == True:
            try:

                email_message.attach(MIMEText(assinatura2, "html"))
                

            except:
                print("Erro na assinatura")

        email_string = email_message.as_string()
        print(email_string)

        context = ssl.create_default_context()
        with smtplib.SMTP(smtp_server, port) as server:
            server.ehlo()  # Can be omitted
            server.starttls(context=context)
            server.ehlo()  # Can be omitted
            server.login(sender_email, password)
            server.sendmail(sender_email, destinatario, email_string)