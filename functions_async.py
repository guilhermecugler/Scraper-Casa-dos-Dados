import asyncio
import json
import math
import os
import time
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen, urlretrieve

import httpx
import pandas
import requests
from bs4 import BeautifulSoup
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
from sheet2dict import Worksheet


def convert_list_to_dict(list):
    dict_from_list = {list[i]: list[i + 1] for i in range(0, len(list), 2)}
    return dict_from_list
    
def get_cnaes():
    response = requests.get(
        f'https://api.casadosdados.com.br/v1/rf/cnpj/query/search/field/cnae',
        headers=HEADERS,
    )


    cnae_name = []
    cnae_code = []

    for cnae in response.json(): 
        cnae_name.append(cnae['name'])
        cnae_code.append(cnae['code'])
    
    
    return cnae_name, cnae_code


def search_city(state):

    response = requests.get(
        f'https://api.casadosdados.com.br/v1/rf/cnpj/query/search/field/municipio/{state}',
        headers=HEADERS,
    )

    cities = []

    for city in response.json(): cities.append(city['name'])

    return cities


list_df_all_cnpj_details = []
list_cnpj_numbers = []
API_CASA_DOS_DADOS = 'https://api.casadosdados.com.br/v2/public/cnpj/search'
URL_DETALHES_CNPJ = 'https://casadosdados.com.br/solucao/cnpj'
HEADERS = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36'
    }


def get_page_count(json_filters):
    pages_count = 1
    try:
        response = requests.post(API_CASA_DOS_DADOS, json=json_filters, headers=HEADERS)
        response_json = response.json()

        cnpj_count = response_json['data']['count']

        # pages_count = math.ceil(cnpj_count / 20) if cnpj_count < 1000 and cnpj_count > 20 else 1

        if cnpj_count < 1000 and cnpj_count > 20 : pages_count = math.ceil(cnpj_count / 20)
        elif cnpj_count : pages_count = 50

    except HTTPError as e:
        print(e.status, e.reason)
    except URLError as e:
        print(e.reason)

    return pages_count


async def search_filters_return_list_cnpj_numbers(json_filters, page_number):
        json_filters.update({'page': page_number})
        async with httpx.AsyncClient() as client:
            req = await client.post(API_CASA_DOS_DADOS, json=json_filters, headers=HEADERS)
            req_json = req.json()['data']['cnpj']
            cnpjs = [cnpj['cnpj'] for cnpj in req_json]
            list_cnpj_numbers.extend(cnpjs)
            return
        
dict_cnpj_details = pandas.DataFrame()

async def get_cnpj_details(client, url):
        response_details = []
        # dict_cnpj_details = {}
        try:

            response = await client.post(url, headers=HEADERS)
            
            # print(response)
            if response.status_code == 200:
                response_details.append(response)
                soup = BeautifulSoup(response.content, 'html.parser')
                cnpj_all_details_raw = soup.find_all('div', class_='column is-narrow')
                cnpj_all_details = []
                dict_cnpj_details = {}

                for i, e in enumerate(cnpj_all_details_raw):
                    cnpj_all_details_raw = soup.find_all('div', class_='column is-narrow')[i]

                    for p in cnpj_all_details_raw.find_all('p'):
                        if len(cnpj_all_details_raw.find_all('p')) == 2:
                            cnpj_all_details.append(p.text)

                    dict_cnpj_details = convert_list_to_dict(cnpj_all_details)

                    municipio = dict_cnpj_details.get('Município')
                    uf = dict_cnpj_details.get('UF')

                    if municipio != None:
                        municipio_formatted = {'Município': municipio.strip()}
                        dict_cnpj_details.update(municipio_formatted)
                        
                    if uf != None:
                        uf_formatted = {'UF': uf.strip()}
                        dict_cnpj_details.update(uf_formatted)


                df_cnpj_details = pandas.DataFrame(data=dict_cnpj_details, index=[1])


                list_df_all_cnpj_details.append(df_cnpj_details)
            else:
                pass
        except HTTPError as e:
            print(e.status, e.reason)

        except URLError as e:
            print(e.reason)
        return list_df_all_cnpj_details



def save_df_list_to_xlsx(file_name, list_df_all_cnpj_details):

    isFile = os.path.isfile(file_name)

    if isFile == False:
        writer = pandas.ExcelWriter(file_name, engine='xlsxwriter')
        writer.close()

    sheet = Worksheet()
    dict_sheet = sheet.xlsx_to_dict(path=file_name)
    df_sheet = pandas.DataFrame(data=dict_sheet.sheet_items)

    try:
        df_sheet_final = pandas.concat(list_df_all_cnpj_details)
        df_sheet_final = pandas.concat([df_sheet, df_sheet_final])
        df_sheet_final.to_excel(file_name, index=False)

    except Exception as e:
        print(e)

    return file_name


def organize_sheet(file_name):

    sheet = Worksheet()

    dict_sheet = sheet.xlsx_to_dict(path=file_name)
    df_sheet = pandas.DataFrame(data=dict_sheet.sheet_items)

    try:
        df_sheet = df_sheet.drop_duplicates(subset='CNPJ')
        df_sheet.to_excel(file_name, index=False)

        writer = pandas.ExcelWriter(
            file_name,
            engine='openpyxl',
            if_sheet_exists='replace',
            mode='a',
        )

        workbook = writer.book
        sheet = workbook['Sheet1']
        #Mudar tamanho das colunas da planilha
        for column_cells in sheet.columns:
            new_column_length = max(
                len(str(cell.value)) for cell in column_cells # type: ignore
            )
            new_column_letter = get_column_letter(column_cells[0].column) # type: ignore
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = (
                    new_column_length * 1.23
                )

        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center'
                )

        writer.close()

        return True
    except PermissionError as pe:
        print('Permission Error')
        return pe.errno

json_filters = {
    'query':{
        'termo':[],
        'atividade_principal':[],
        'natureza_juridica':[],
        'uf':[],
        'municipio':[],
        'situacao_cadastral':'ATIVA',
        'cep':[],
        'ddd':[]
    },
    'range_query':{
        'data_abertura':{
            'lte':None,
            'gte':None
        },
        'capital_social':{
            'lte':None,
            'gte':None
        }
    },
    'extras':{
        'somente_mei':False,
        'excluir_mei':False,
        'com_email':False,
        'incluir_atividade_secundaria':False,
        'com_contato_telefonico':False,
        'somente_fixo':False,
        'somente_celular':False,
        'somente_matriz':False,
        'somente_filial':False
    },
    'page':1
}


async def get_list_cnpj_numbers():
    print('Pegando lista de cnpj')
    # pages_count = get_page_count(json_filters)
    pages_count = 3
    tasks = []
    for page in range(pages_count):
        tasks.append(search_filters_return_list_cnpj_numbers(json_filters, page))
    await asyncio.gather(*tasks)


async def main():
    # pages_count = get_page_count(json_filters)
    print('Iniciando busca de detalhes')

    async with httpx.AsyncClient(follow_redirects = True, timeout=None) as client:

        tasks = []

        for cnpj in list_cnpj_numbers:
            url = f'{URL_DETALHES_CNPJ}/{cnpj}'
            tasks.append(asyncio.ensure_future(get_cnpj_details(client, url)))

        results = await asyncio.gather(*tasks)
        print(len(results))

        # for cnpj in cnpj_t:
        #     print(cnpj)

# asyncio.run(get_list_cnpj_numbers())

# asyncio.run(main())


# save_df_list_to_xlsx(file_name='ResultadoAsync.xlsx', list_df_all_cnpj_details=list_df_all_cnpj_details)
# organize_sheet(file_name='ResultadoAsync.xlsx')
# print("--- %s seconds ---" % (time.time() - start_time))