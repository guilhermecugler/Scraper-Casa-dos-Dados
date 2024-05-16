#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Description: Scraper com interface utilizando CustomTkinter do site casadosdados.com.br
Author: Guilherme Cugler
"""

import asyncio
import json
import math
import os
import time
import tkinter
from datetime import datetime
from threading import Thread, enumerate as enumt, Event
from urllib.error import HTTPError, URLError
import customtkinter
import httpx
import pandas
import requests
from bs4 import BeautifulSoup
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
from PIL import Image
from sheet2dict import Worksheet
from tkinter import filedialog as fd
from aiohttp import ClientSession, TCPConnector, ClientTimeout

stop_flag = False
MAX_TASKS = 10
global stop_event
global bloqueio_502

bloqueio_502 = Event()
sessions = []

stop_event = Event()

def start_thread(function_name):
    print(f'Thread {function_name} started')

    global stop
    stop = 0

    t = Thread(target=function_name)
    t.daemon = True
    t.start()


global list_df_all_cnpj_details
list_df_all_cnpj_details = []

global progress_step
progress_step = 0


global iter_step

class Functions:
    def __init__(self, master):
        super().__init__()
        
    list_cnpj_numbers = []
    API_CASA_DOS_DADOS = 'https://api.casadosdados.com.br/v2/public/cnpj/search'
    URL_DETALHES_CNPJ = 'https://casadosdados.com.br/solucao/cnpj'
    HEADERS = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36'
        }
    
    def get_tasks(self):
        tasks = asyncio.all_tasks()
        # report all tasks
        for task in tasks:
            print(f'> {task.get_name()}, {task.get_coro()}')

    def convert_list_to_dict(self, list):
        dict_from_list = {list[i]: list[i + 1] for i in range(0, len(list), 2)}
        return dict_from_list
        
    def get_cnaes(self):
        response = requests.get(
            f'https://api.casadosdados.com.br/v1/rf/cnpj/query/search/field/cnae',
            headers=self.HEADERS,
        )


        cnae_name = []
        cnae_code = []

        for cnae in response.json():
            cnae_name.append(cnae['name'])
            cnae_code.append(cnae['code'])
        
        
        return cnae_name, cnae_code


    def search_city(self, state):

        response = requests.get(
            f'https://api.casadosdados.com.br/v1/rf/cnpj/query/search/field/municipio/{state}',
            headers=self.HEADERS,
        )

        cities = []

        for city in response.json(): cities.append(city['name'])
        cities.sort()

        return cities


    def get_page_count(self, json_filters):
        pages_count = 1
        try:
            response = requests.post(self.API_CASA_DOS_DADOS, json=json_filters, headers=self.HEADERS)
            response_json = response.json()

            cnpj_count = response_json['data']['count']
            # pages_count = math.ceil(cnpj_count / 20) if cnpj_count < 1000 and cnpj_count > 20 else 1

            if cnpj_count < 1000 and cnpj_count > 20 : pages_count = math.ceil(cnpj_count / 20)
            elif cnpj_count > 1000 : pages_count = 50
            else: pages_count = 1

            # print(f'quantidade cnpj {cnpj_count} paginas {pages_count}')

        except HTTPError as e:
            print(e.status, e.reason)
        except URLError as e:
            print(e.reason)
        except KeyError as e:
            print('Nenhuma página com CNPJ encontrada')
            return 0

        return pages_count
    
    def get_save_folder(self):

        directory = fd.askdirectory(
            title='Selecionar pasta'
        )
        return directory

    async def search_filters_return_list_cnpj_numbers(self, json_filters, page_number, client):
            if stop_event.is_set():
                return 
            json_filters.update({'page': page_number})
            try:
                # async with httpx.AsyncClient() as client:
                    req = await client.post(self.API_CASA_DOS_DADOS, json=json_filters, headers=self.HEADERS)
                    
                    res = await req.text()

                    req_json = json.loads(res)['data']['cnpj']
                    cnpjs = [cnpj['cnpj'] for cnpj in req_json]
                    self.list_cnpj_numbers.extend(cnpjs)
                    
                    return 1
            except KeyError as e:
                print('Nenhum CNPJ encontrado')
                return 0
            except json.decoder.JSONDecodeError as e:
                return "JSONDecodeError"


    async def get_cnpj_details(self, url):
            if stop_event.is_set():
                return 
                
            while not sessions:
                await asyncio.sleep(1)

            response_details = []
            # dict_cnpj_details = {}
            try:
                # async with httpx.AsyncClient(follow_redirects = True, timeout=None) as session: # httpx
                # async with ClientSession() as session:
                async with sessions[-1].get(url) as response:
                    if response.status == 200:

                        html = await response.text()
                    
                    # print(response)
                    # if response.status_code == 200: # httpx

                        response_details.append(response)
                        soup = BeautifulSoup(html, 'html.parser')
                        # soup = BeautifulSoup(response.content, 'html.parser') # httpx

                        # cnpj_all_details_raw = soup.find_all('div', class_='column is-narrow')
                        cnpj_all_details = []
                        dict_cnpj_details = {}
                        cnpj_all_details_raw = soup.find_all('label')

                        for i, label in enumerate(cnpj_all_details_raw):
                            if i == 0:
                                continue
                            if label.find_next().name != "p":
                                continue
                            dict_cnpj_details[label.text.replace(":", "")] = label.find_next("p").text

                        df_cnpj_details = pandas.DataFrame(data=dict_cnpj_details, index=[1])
                        global progress_step
                        progress_step += iter_step
                        app.progress_bar_update(progress_step)

                        list_df_all_cnpj_details.append(df_cnpj_details)

                    elif response.status == 502:
                        bloqueio_502.set()
                        print(await response.text())
                        progress_step += iter_step
                        app.progress_bar_update(progress_step)
                    
            except HTTPError as e:
                print(e.status, e.reason)

            except URLError as e:
                print(e.reason)

            except httpx.ReadError:
                pass
            except Exception as e:
                print(e)
                pass


    def save_df_list_to_xlsx(self, file_name, list_df_all_cnpj_details):
        app.status_update(f"Criando planilha...")

        isFile = os.path.isfile(file_name)

        if isFile == False:
            writer = pandas.ExcelWriter(file_name, engine='xlsxwriter')
            writer.close()

        sheet = Worksheet()
        dict_sheet = sheet.xlsx_to_dict(path=file_name)
        df_sheet = pandas.DataFrame(data=dict_sheet.sheet_items)

        try:
            df_sheet_final = pandas.concat(list_df_all_cnpj_details)
            df_sheet_final = pandas.concat([df_sheet, df_sheet_final])
            df_sheet_final.to_excel(file_name, index=False)

        except Exception as e:
            print(e)

        return file_name


    def organize_sheet(self, file_name):

        app.status_update(f"Organizando planilha, aguarde!")
        

        sheet = Worksheet()

        dict_sheet = sheet.xlsx_to_dict(path=file_name)
        df_sheet = pandas.DataFrame(data=dict_sheet.sheet_items)

        try:
            df_sheet = df_sheet.drop_duplicates(subset='CNPJ')

            global n_cnpj
            n_cnpj = len(df_sheet)

            df_sheet.to_excel(file_name, index=False)

            writer = pandas.ExcelWriter(
                file_name,
                engine='openpyxl',
                if_sheet_exists='replace',
                mode='a',
            )

            workbook = writer.book
            sheet = workbook['Sheet1']
            #Mudar tamanho das colunas da planilha
            for column_cells in sheet.columns: # type: ignore
                new_column_length = max(
                    len(str(cell.value)) for cell in column_cells # type: ignore
                )
                new_column_letter = get_column_letter(column_cells[0].column) # type: ignore
                if new_column_length > 0:
                    sheet.column_dimensions[new_column_letter].width = ( # type: ignore
                        new_column_length * 1.23
                    )

            for row in range(1, sheet.max_row + 1): # type: ignore
                for col in range(1, sheet.max_column + 1): # type: ignore
                    cell = sheet.cell(row, col) # type: ignore
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center'
                    )

            writer.close()

            return True
        except PermissionError as pe:
            app.status_update(f"Feche a planilha e tente novamente!")
            print('Permission Error')
            return pe.errno

    json_filters = {
        'query':{
            'termo':[],
            'atividade_principal':[],
            'natureza_juridica':[],
            'uf':[],
            'municipio':[],
            'situacao_cadastral':'ATIVA',
            'cep':[],
            'ddd':[]
        },
        'range_query':{
            'data_abertura':{
                'lte':None,
                'gte':None
            },
            'capital_social':{
                'lte':None,
                'gte':None
            }
        },
        'extras':{
            'somente_mei':False,
            'excluir_mei':False,
            'com_email':False,
            'incluir_atividade_secundaria':False,
            'com_contato_telefonico':False,
            'somente_fixo':False,
            'somente_celular':False,
            'somente_matriz':False,
            'somente_filial':False
        },
        'page':1
    }

######################## TESTE LIMITAR REQUESTS #######################
    # async def get_list_cnpj_numbers(self):

        
    #     if stop_event.is_set():
    #         return 
    #     print('Pegando lista de cnpj')


    #     pages_count = self.get_page_count(self.json_filters)


    #     print(f'numero de paginas: {pages_count}')

    #     if pages_count == 0:
    #         print(pages_count)
    #         return 0
    #     async with ClientSession(connector=TCPConnector(limit=10, limit_per_host=10)) as client:
    #                 conc_req = 40

    #                 await self.gather_with_concurrency(conc_req, *[self.search_filters_return_list_cnpj_numbers(self.json_filters, page, client) for page in range(pages_count)])

######################## TESTE LIMITAR REQUESTS #######################



    async def refresh_session(self):
        # this function periodically refreshes the token every X sec
        connector = TCPConnector(limit=3)
        timeout = ClientTimeout(
            total=None, connect=300, sock_connect=300, sock_read=None
        )

        while True:

            sessions.append(
                ClientSession(headers=self.HEADERS)
            )
            # print("New session created")
            await asyncio.sleep(5)  # every 5 seconds refresh session

    async def get_list_cnpj_numbers(self):

        
        if stop_event.is_set():
            return 
        print('Pegando lista de cnpj')


        pages_count = self.get_page_count(self.json_filters)


        print(f'numero de paginas: {pages_count}')

        if pages_count == 0:
            print(pages_count)
            return 0
        # pages_count = 3
        tasks = []
        async with ClientSession() as client:
            for page in range(pages_count):
                if stop_event.is_set():
                    return 
                tasks.append(self.search_filters_return_list_cnpj_numbers(self.json_filters, page, client))
                
                # tasks.append(asyncio.ensure_future(self.search_filters_return_list_cnpj_numbers(self.json_filters, page, client)))
            if not stop_flag:
                await asyncio.gather(*tasks)

    async def gather_with_concurrency(self, n, *tasks):
        semaphore = asyncio.Semaphore(n)

        async def sem_task(task):
            async with semaphore:
                return await task

        return await asyncio.gather(*(sem_task(task) for task in tasks))

    async def main(self):
        if stop_flag: return

        # pages_count = get_page_count(json_filters)
        print('Iniciando busca de detalhes')
        self.list_cnpj_numbers = list(dict.fromkeys(self.list_cnpj_numbers))
        print(f'Buscando {len(self.list_cnpj_numbers)} CNPJ(s)... aguarde')

        app.status_update(text=f'Buscando {len(self.list_cnpj_numbers)} CNPJ(s)... aguarde')

        if self.list_cnpj_numbers == 0:
            app.status_update(text='Nenhuma empresa encontrada')

            print('Nenhuma empresa encontrada')
            return 

        if self.list_cnpj_numbers == []:
            return
        
        global iter_step
        iter_step  = 1 / len(self.list_cnpj_numbers)

        global progress_step
        progress_step = iter_step

        # async with httpx.AsyncClient(follow_redirects = True, timeout=None) as client:
        # async with httpx.AsyncClient(follow_redirects = True, timeout=None) as client:
        conc_req = 40


        t1 = asyncio.create_task(self.refresh_session())
    
        t2 = asyncio.create_task(self.gather_with_concurrency(conc_req, *[self.get_cnpj_details(f'{self.URL_DETALHES_CNPJ}/{cnpj}') for cnpj in self.list_cnpj_numbers]))

            # ret = await asyncio.gather(*[self.get_cnpj_details(f'{self.URL_DETALHES_CNPJ}/{cnpj}', client) for cnpj in self.list_cnpj_numbers])
        
        done, _ = await asyncio.wait([t1, t2], return_when=asyncio.FIRST_COMPLETED)

        for s in sessions:
            await s.close()

        # await self.gather_with_concurrency(conc_req, *[self.get_cnpj_details(f'{self.URL_DETALHES_CNPJ}/{cnpj}') for cnpj in self.list_cnpj_numbers])



class FiltersFrame(customtkinter.CTkFrame):
    def __init__(self, master, title):
        super().__init__(master)
        self.grid_columnconfigure(2, weight=1)
        self.title = title
        self.checkboxes = []


        self.title = customtkinter.CTkLabel(self, text=self.title, corner_radius=6)
        self.title.grid(row=0, column=0, padx=10, pady=(10, 0), sticky="ew", columnspan=3)

        functions = Functions(self)
        cnaes = functions.get_cnaes()

        def combobox_estados_callback(choice):
                    self.combobox_estados_var.set(choice)
                    self.combobox_municipios_var.set('Todos Municipios')
                    self.combobox_municipios.configure(values=functions.search_city(self.combobox_estados_var.get()))

        def combobox_municipios_callback(choice):
            self.combobox_municipios_var.set(choice)    

        def combobox_cnae_callback(choice):
            for i, cnae in enumerate(cnaes[0]):
                if cnae == choice:
                    self.cnae_code_var.set(cnaes[1][i])
                    print(f"cnae selected code: {self.cnae_code_var.get()}")

        def format_date_inicial(event):
            
            text = self.entry_data_inicial.get().replace("/", "")[:8]
            new_text = ""

            if event.keysym.lower() == "backspace": return
            
            for index in range(len(text)):
                if not text[index] in "0123456789": continue
                if index in [1, 3]: new_text += text[index] + "/"
                else: new_text += text[index]

            self.entry_data_inicial.delete(0, "end")
            self.entry_data_inicial.insert(0, new_text)

        def format_date_final(event):
            
            text = self.entry_data_final.get().replace("/", "")[:8]
            new_text = ""

            if event.keysym.lower() == "backspace": return
            
            for index in range(len(text)):
                if not text[index] in "0123456789": continue
                if index in [1, 3]: new_text += text[index] + "/"
                else: new_text += text[index]

            self.entry_data_final.delete(0, "end")
            self.entry_data_final.insert(0, new_text)

        self.check_somente_mei_var = customtkinter.BooleanVar(value=False)
        self.check_excluir_mei_var = customtkinter.BooleanVar(value=False)
        self.check_com_telefone_var = customtkinter.BooleanVar(value=False)
        self.check_somente_fixo_var = customtkinter.BooleanVar(value=False)
        self.check_somente_matriz_var = customtkinter.BooleanVar(value=False)
        self.check_somente_filial_var = customtkinter.BooleanVar(value=False)
        self.check_somente_celular_var = customtkinter.BooleanVar(value=False)
        self.check_com_email_var = customtkinter.BooleanVar(value=False)
        self.check_atividade_secundaria_var = customtkinter.BooleanVar(value=False)
        self.combobox_estados_var = customtkinter.StringVar(value='Todos Estados')
        self.combobox_municipios_var = customtkinter.StringVar(value='Todos Municipios')
        self.combobox_cnae_var = customtkinter.StringVar(value='Todas Atividades')
        self.cnae_code_var = customtkinter.StringVar(value='')

        self.entry_termo = customtkinter.CTkEntry(
            self,
            placeholder_text='Razão Social ou Termo - Ex: Celular'
            )
        self.entry_termo.grid(row=1, columnspan=3, column=0, padx=10, pady=10, sticky='ew')

        self.check_somente_mei = customtkinter.CTkCheckBox(
            self,
            text='Somente MEI',
            variable=self.check_somente_mei_var,
            onvalue=True,
            offvalue=False,
        )
        self.check_somente_mei.grid(row=2, column=0, padx=10, pady=10, sticky='ew')

        self.check_excluir_mei = customtkinter.CTkCheckBox(
            self,
            text='Excluir MEI',
            variable=self.check_excluir_mei_var,
            onvalue=True,
            offvalue=False,
        )
        self.check_excluir_mei.grid(row=2, column=1, padx=10, pady=10, sticky='ew')

        self.check_com_telefone = customtkinter.CTkCheckBox(
            self,
            text='Com Telefone',
            variable=self.check_com_telefone_var,
            onvalue=True,
            offvalue=False,
        )
        self.check_com_telefone.grid(row=2, column=2, padx=10, pady=10, sticky='ew')

        self.check_somente_fixo = customtkinter.CTkCheckBox(
            self,
            text='Somente Fixo',
            variable=self.check_somente_fixo_var,
            onvalue=True,
            offvalue=False,
        )
        self.check_somente_fixo.grid(row=3, column=0, padx=10, pady=10, sticky='ew')

        self.check_somente_matriz = customtkinter.CTkCheckBox(
            self,
            text='Somente Matriz',
            variable=self.check_somente_matriz_var,
            onvalue=True,
            offvalue=False,
        )
        self.check_somente_matriz.grid(row=3, column=1, padx=10, pady=10, sticky='ew')

        self.check_somente_filial = customtkinter.CTkCheckBox(
            self,
            text='Somente Filial',
            variable=self.check_somente_filial_var,
            onvalue=True,
            offvalue=False,
        )
        self.check_somente_filial.grid(row=3, column=2, padx=10, pady=10, sticky='ew')

        self.check_somente_celular = customtkinter.CTkCheckBox(
            self,
            text='Somente Celular',
            variable=self.check_somente_celular_var,
            onvalue=True,
            offvalue=False,
        )
        self.check_somente_celular.grid(row=4, column=0, padx=10, pady=10, sticky='ew')

        self.check_com_email = customtkinter.CTkCheckBox(
            self,
            text='Com E-mail',
            variable=self.check_com_email_var,
            onvalue=True,
            offvalue=False,
        )
        self.check_com_email.grid(row=4, column=1, padx=10, pady=10, sticky='ew')

        self.check_atividade_secundaria = customtkinter.CTkCheckBox(
            self,
            text='Atividade Secundária',
            variable=self.check_atividade_secundaria_var,
            onvalue=True,
            offvalue=False,
        )
        self.check_atividade_secundaria.grid(
            row=4, column=2, padx=10, pady=10, sticky='ew'
        )

        self.combobox_estados = customtkinter.CTkComboBox(
            self, values = ['Todos Estados','AC','AL','AP','AM','BA','CE','DF','ES','GO','MA','MS','MT','MG','PA','PB','PR','PE','PI','RJ','RN','RS','RO','RR','SC','SP','SE','TO',],
            command=combobox_estados_callback,
            variable=self.combobox_estados_var,
        )
        self.combobox_estados.grid(row=5, column=0, padx=10, pady=10, sticky='ew')

        # self.combobox_municipios = CTkScrollableDropdown(
        #     self,
        #     values=[''],
        #     command=combobox_municipios_callback,
        #     variable=self.combobox_municipios_var,
        # )
        # self.combobox_municipios.grid(row=5, column=1, padx=10, pady=10, sticky='ew')
        self.combobox_municipios = customtkinter.CTkComboBox(
            self,
            values=[''],
            command=combobox_municipios_callback,
            variable=self.combobox_municipios_var,
        )
        self.combobox_municipios.grid(row=5, column=1, padx=10, pady=10, sticky='ew')

        self.combobox_cnaes = customtkinter.CTkComboBox(
            self,
            values=[''],
            command=combobox_cnae_callback,
            variable=self.combobox_cnae_var)
      
        self.entry_bairro = customtkinter.CTkEntry(self, placeholder_text='Bairro')
        self.entry_bairro.grid(row=5, column=2, padx=10, pady=10, sticky='ew')

        self.entry_CEP = customtkinter.CTkEntry(self, placeholder_text='CEP')
        self.entry_CEP.grid(row=7, column=1, padx=10, pady=10, sticky='ew')

        self.entry_DDD = customtkinter.CTkEntry(self, placeholder_text='DDD')
        self.entry_DDD.grid(row=7, column=2, padx=10, pady=10, sticky='ew')

        self.combobox_cnaes.grid(row=7, column=0, padx=10, pady=10, sticky='ew')
        self.combobox_cnaes.configure(values=cnaes[0])

        self.label_data = customtkinter.CTkLabel(self, text="Período de abertura")
        self.label_data.grid(row=8, column=0, padx=10, pady=10, sticky='ew')


        self.entry_data_inicial = customtkinter.CTkEntry(self, placeholder_text='Inicio - 01/01/2023')
        self.entry_data_inicial.bind("<KeyRelease>", command=format_date_inicial)
        self.entry_data_inicial.grid(row=8, column=1, padx=10, pady=10, sticky='ew')
        

        self.entry_data_final = customtkinter.CTkEntry(self, placeholder_text='Fim - 01/12/2023')
        self.entry_data_final.bind("<KeyRelease>", command=format_date_final)
        self.entry_data_final.grid(row=8, column=2, padx=10, pady=10, sticky='ew')

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.width = 550   # Width
        self.height = 650   # Height

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight() 

        x = (screen_width / 2) - (self.width / 2)
        y = (screen_height / 2) - (self.height / 2)

        self.title("Casa dos Dados")
        self.geometry('%dx%d+%d+%d' % (self.width, self.height, x, y-30))
        self.grid_columnconfigure((0, 1), weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.resizable(width=False, height=False)

        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "images")

        self.iconbitmap(os.path.join(image_path, "icon.ico"))
        
        self.radio_var = tkinter.IntVar(value=0)


        self.home_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(image_path, "logo_casa_dos_dados_light.png")),
                                                 dark_image=Image.open(os.path.join(image_path, "logo_casa_dos_dados_dark.png")), size=(500, 56))
        self.home_image_label = customtkinter.CTkLabel(self, text="", image=self.home_image)
        self.home_image_label.grid(row=0, column=0, padx=10, pady=(30, 10), sticky="ew", columnspan=4)

        self.filters_frame = FiltersFrame(self, "Filtros")
        self.filters_frame.grid(row=1, column=0, padx=10, pady=(10, 0), sticky="nsew", columnspan=4)


        self.label_numero_buscas = customtkinter.CTkLabel(self, text="Repetir")
        self.label_numero_buscas.grid(row=5, column=0, padx=(20, 0), pady=10, sticky="w")

        self.button_diminuir_buscas = customtkinter.CTkButton(self, text="-", width=20, command=self.button_diminuir_buscas_callback)
        self.button_diminuir_buscas.grid(row=5, column=0, padx=(70, 0), pady=10, sticky="w")

        self.entry_quantidade_buscas_var = customtkinter.IntVar()
        self.entry_quantidade_buscas_var.set(1)
        self.entry_quantidade_buscas = customtkinter.CTkEntry(self, placeholder_text="1", width=30, textvariable=self.entry_quantidade_buscas_var)
        self.entry_quantidade_buscas.grid(row=5, column=0, padx=(0, 23), pady=10, sticky="e")

        self.button_aumentar_buscas = customtkinter.CTkButton(self, text="+", width=20, command=self.button_aumentar_buscas_callback)
        self.button_aumentar_buscas.grid(row=5, column=0, padx=0, pady=10, sticky="e")

        
        self.button_buscar_empresas = customtkinter.CTkButton(self, text="Buscar Empresas", command=self.button_buscar_empresas_callback)
        self.button_buscar_empresas.grid(row=5, column=1, padx=10, pady=10, sticky="ew", columnspan=2)

        self.button_cancelar = customtkinter.CTkButton(self, text="Cancelar", command=self.button_cancelar_callback, state="disabled")
        self.button_cancelar.grid(row=5, column=3, padx=10, pady=10, sticky="ew", columnspan=1)

        self.status = customtkinter.CTkLabel(self, text="Faça uma busca!")
        self.status.grid(row=6, column=0, padx=0, pady=0, sticky="ew", columnspan=4)

        self.progress_bar = customtkinter.CTkProgressBar(self, orientation='horizontal')
        self.progress_bar.grid(row=7, column=0, padx=10, pady=(5, 5), sticky="ew", columnspan=4)
        self.progress_bar.set(0)

        self.label_file_type = customtkinter.CTkLabel(self, text="Tipo de arquivo:")
        self.label_file_type.grid(row=8, column=1)

        self.file_type_var = customtkinter.StringVar(value='xlsx')

        self.radio_xlsx = customtkinter.CTkRadioButton(self, text='Planilha', value='xlsx', variable=self.file_type_var, command=self.radiobutton_event, radiobutton_width=13, radiobutton_height=13)
        self.radio_xlsx.grid(row=8, column=2, padx=0, pady=0, sticky="w")

        self.radio_csv = customtkinter.CTkRadioButton(self, text='CSV', value='csv', variable=self.file_type_var, command=self.radiobutton_event, radiobutton_width=13, radiobutton_height=13)
        self.radio_csv.grid(row=8, column=3, padx=0, pady=0, sticky="e")




        self.file_entry_var = customtkinter.Variable(value=f"{datetime.strftime(datetime.now(), '%d-%m-%Y %H-%M')}.{self.file_type_var.get()}")


        self.file_entry = customtkinter.CTkEntry(self, textvariable=self.file_entry_var)
        self.file_entry.grid(row=9, column=1, padx=10, pady=20, sticky="ew")

        self.button_select_folder = customtkinter.CTkButton(self, text="Selecionar Pasta", command=self.button_select_folder_callback)
        self.button_select_folder.grid(row=9, column=2, padx=10, pady=10, sticky="ew", columnspan=2)

        
        self.appearance_mode_menu = customtkinter.CTkOptionMenu(self, values=["Sistema", "Escuro", "Claro"],
                                                                command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=9, column=0, padx=10, pady=20, sticky="ws")

    


    def button_select_folder_callback(self):
        functions = Functions(self)
        directory = functions.get_save_folder()
        if directory == '':
            return
        file_location = f"{directory}/{datetime.strftime(datetime.now(), '%d-%m-%Y %H-%M')}.{self.file_type_var.get()}"
        self.file_entry_var.set(file_location.replace('//', '/'))

    def radiobutton_event(self):
        self.file_entry_var.set(f"{datetime.strftime(datetime.now(), '%d-%m-%Y %H-%M')}.{self.file_type_var.get()}")

    def progress_bar_update(self, step):
        self.progress_bar.set(step)
        self.update_idletasks()

    def status_update(self, text):
        self.status.configure(text=text)
        self.update_idletasks()

    def button_cancelar_callback(self):
        functions = Functions(self)
        # self.button_cancelar.configure(state='normal')

        stop_event.set()
        # self.button_buscar_empresas.configure(state='normal')
        self.button_cancelar.configure(state='disabled')

        app.status_update("Cancelando, aguarde...")
        # report all tasks
        # global stop_flag
        # stop_flag = True

        # tasks = asyncio.all_tasks()
        # print("CLICOU CANCELAR")
        # for task in tasks:
            # print(task.get_name)

    def button_aumentar_buscas_callback(self):
        novo_valor = self.entry_quantidade_buscas_var.get()+1
        self.entry_quantidade_buscas_var.set(novo_valor)

    def button_diminuir_buscas_callback(self):
        novo_valor = self.entry_quantidade_buscas_var.get()-1
        if novo_valor < 1:
            self.entry_quantidade_buscas_var.set(1)
        else:
            self.entry_quantidade_buscas_var.set(novo_valor)

    def button_buscar_empresas_callback(self):
        self.button_buscar_empresas.configure(state='disabled')
        self.button_cancelar.configure(state='normal')
        
        stop_event.clear()

        functions = Functions(self)

        try:
            functions.json_filters.update(
                {
            'query': {
                'termo': [] if self.filters_frame.entry_termo.get() == '' else [self.filters_frame.entry_termo.get()],
                'atividade_principal': [] if self.filters_frame.cnae_code_var.get() == '' else [self.filters_frame.cnae_code_var.get()],
                'natureza_juridica': [],
                'uf': [] if self.filters_frame.combobox_estados_var.get() == 'Todos Estados' else [self.filters_frame.combobox_estados_var.get()],
                'municipio': [] if self.filters_frame.combobox_municipios_var.get() == 'Todos Municipios' else [self.filters_frame.combobox_municipios_var.get()],
                'cep': [] if self.filters_frame.entry_CEP.get() == '' else [self.filters_frame.entry_CEP.get()],
                'ddd': [] if self.filters_frame.entry_DDD.get() == '' else [self.filters_frame.entry_DDD.get()],
                }, 
            'range_query':{
                'data_abertura':{
                    'lte': None if self.filters_frame.entry_data_final.get() == '' else datetime.strftime(datetime.strptime(self.filters_frame.entry_data_final.get(), '%d/%m/%Y'), '%Y-%m-%d'),
                    'gte': None if self.filters_frame.entry_data_inicial.get() == '' else datetime.strftime(datetime.strptime(self.filters_frame.entry_data_inicial.get(), '%d/%m/%Y'), '%Y-%m-%d'),
                }
            },
            'extras':{
            'somente_mei':self.filters_frame.check_somente_mei_var.get(),
            'excluir_mei':self.filters_frame.check_excluir_mei_var.get(),
            'com_email':self.filters_frame.check_com_email_var.get(),
            'incluir_atividade_secundaria':self.filters_frame.check_atividade_secundaria_var.get(),
            'com_contato_telefonico':self.filters_frame.check_somente_fixo_var.get(),
            'somente_fixo':self.filters_frame.check_somente_fixo_var.get(),
            'somente_celular':self.filters_frame.check_somente_celular_var.get(),
            'somente_matriz':self.filters_frame.check_somente_matriz_var.get(),
            'somente_filial':self.filters_frame.check_somente_filial_var.get()
        },})
            
        except ValueError as e:
            print(f'Error: {e}')
            self.button_buscar_empresas.configure(state='normal')
            self.button_cancelar.configure(state='disabled')

            return

        self.progress_bar.set(0)
        # print(functions.json_filters)
        
        def buscar():
            list_df_all_cnpj_details.clear()

            
            app.status_update(text=f"Iniciando módulo de busca... aguarde!")

            


            global progress_step
            global iter_step

            repetir = self.entry_quantidade_buscas_var.get()

            self.progress_bar.configure(mode="indeterminate")

            self.progress_bar.start()
            
            for i in (range(repetir)):
                app.status_update(text=f"Calculando quantidade aproximada de CNPJs...({len(functions.list_cnpj_numbers)})")
                get_cnpjs__list = asyncio.run(functions.get_list_cnpj_numbers())
                functions.list_cnpj_numbers = list(dict.fromkeys(functions.list_cnpj_numbers))

            self.progress_bar.stop()
            self.progress_bar.configure(mode="determinate")
            

            
            start_time = time.time()


            if get_cnpjs__list == 0:
                list_df_all_cnpj_details.clear()
                self.button_buscar_empresas.configure(state='normal')
                self.button_cancelar.configure(state='disabled')

                # print("--- %s seconds ---" % (time.time() - start_time))
                return
            
            app.status_update(text=f"Encontrados aproximadamente {len(functions.list_cnpj_numbers)} CNPJ(s), iniciando extração...")

         

            asyncio.run(functions.main())


            file_name = self.file_entry_var.get()
            # print(file_name)

            if self.file_type_var.get() == 'xlsx':
                functions.save_df_list_to_xlsx(file_name, list_df_all_cnpj_details)
                functions.organize_sheet(file_name)
            else:
                pandas.concat(list_df_all_cnpj_details).to_csv(file_name, index=False)

            app.status_update(f"Finalizado... salvos {n_cnpj} CNPJ(s)")
            
            list_df_all_cnpj_details.clear()
            functions.list_cnpj_numbers.clear()
            
            self.button_buscar_empresas.configure(state='normal')
            self.button_cancelar.configure(state='disabled')
            
            print("--- %s seconds ---" % (time.time() - start_time))
            teste = enumt()
            # print(teste)
        

        start_thread(buscar)

        teste = enumt()
        # print(teste)
        


    def button_callback2(self):
        functions = Functions(self)

        functions.json_filters.update(
            {
        'query': {
            'termo': [] if self.filters_frame.entry_termo.get() == '' else [self.filters_frame.entry_termo.get()],
            'atividade_principal': [] if self.filters_frame.cnae_code_var.get() == '' else [self.filters_frame.cnae_code_var.get()],
            'natureza_juridica': [],
            'uf': [] if self.filters_frame.combobox_estados_var.get() == 'Todos Estados' else [self.filters_frame.combobox_estados_var.get()],
            'municipio': [] if self.filters_frame.combobox_municipios_var.get() == 'Todos Municipios' else [self.filters_frame.combobox_municipios_var.get()],
            'situacao_cadastral':'ATIVA',
            'cep': [] if self.filters_frame.entry_CEP.get() == '' else [self.filters_frame.entry_CEP.get()],
            'ddd': [] if self.filters_frame.entry_DDD.get() == '' else [self.filters_frame.entry_DDD.get()],
            }, 
        'range_query':{
            'data_abertura':{
                'lte': None if self.filters_frame.entry_data_inicial.get() == '' else datetime.strftime(datetime.strptime(self.filters_frame.entry_data_inicial.get(), '%d/%m/%Y'), '%d/%m/%Y'),
                'gte': None if self.filters_frame.entry_data_final.get() == '' else datetime.strftime(datetime.strptime(self.filters_frame.entry_data_final.get(), '%d/%m/%Y'), '%d/%m/%Y'),
            }
        },
        'extras':{
        'somente_mei':self.filters_frame.check_somente_mei_var.get(),
        'excluir_mei':self.filters_frame.check_excluir_mei_var.get(),
        'com_email':self.filters_frame.check_com_email_var.get(),
        'incluir_atividade_secundaria':self.filters_frame.check_atividade_secundaria_var.get(),
        'com_contato_telefonico':self.filters_frame.check_somente_fixo_var.get(),
        'somente_fixo':self.filters_frame.check_somente_fixo_var.get(),
        'somente_celular':self.filters_frame.check_somente_celular_var.get(),
        'somente_matriz':self.filters_frame.check_somente_matriz_var.get(),
        'somente_filial':self.filters_frame.check_somente_filial_var.get()
    },})
        
        # print(functions.json_filters)
        
    def change_appearance_mode_event(self, new_appearance_mode):
        if new_appearance_mode == "Escuro": new_appearance_mode = "Dark"
        if new_appearance_mode == "Claro": new_appearance_mode = "Light"
        if new_appearance_mode == "Sistema": new_appearance_mode = "System"

        customtkinter.set_appearance_mode(new_appearance_mode)

app = App()
app.change_appearance_mode_event('System')
app.mainloop()